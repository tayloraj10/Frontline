import csv
import io
import re
from datetime import datetime, timezone
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from pydantic import BaseModel, field_validator, model_validator
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.database import get_db
from app.services.admin_roles import has_admin_role
from app.services.event_permissions import is_group_admin as _is_group_admin
from app.services.stats_window import resolve_stats_window, trend_bucket_unit

router = APIRouter(prefix="/team-events", tags=["team-events"])

SLUG_RE = re.compile(r"^[a-z0-9]+(-[a-z0-9]+)*$")


async def _can_manage_team_event(db: AsyncSession, team_event_id: UUID, user_id: UUID) -> bool:
    if await has_admin_role(db, user_id, "event_manager"):
        return True
    row = await db.execute(
        text("SELECT 1 FROM team_event_organizers WHERE team_event_id = :id AND user_id = :user_id"),
        {"id": str(team_event_id), "user_id": str(user_id)},
    )
    return row.fetchone() is not None


async def _get_event_or_404(db: AsyncSession, team_event_id: UUID):
    result = await db.execute(
        text("""
            SELECT id, campaign_id, slug, title, description, status, starts_at, ends_at,
                   submission_mode, requires_photo, image_url, created_by
            FROM team_events WHERE id = :id
        """),
        {"id": str(team_event_id)},
    )
    row = result.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Event not found")
    return row


class TeamInput(BaseModel):
    name: str
    color: str | None = None
    logo_url: str | None = None


class CreateTeamEventRequest(BaseModel):
    requesting_user_id: UUID
    campaign_id: UUID | None = None
    slug: str
    title: str
    description: str | None = None
    starts_at: datetime
    ends_at: datetime | None = None
    submission_mode: str = "manual_opt_in"
    requires_photo: bool = True
    image_url: str | None = None
    teams: list[TeamInput]

    @field_validator("slug")
    @classmethod
    def _valid_slug(cls, v: str) -> str:
        if not SLUG_RE.match(v):
            raise ValueError("slug must be lowercase alphanumeric with single hyphens")
        return v

    @field_validator("submission_mode")
    @classmethod
    def _valid_mode(cls, v: str) -> str:
        if v not in ("automatic", "manual_opt_in"):
            raise ValueError("submission_mode must be 'automatic' or 'manual_opt_in'")
        return v

    @field_validator("teams")
    @classmethod
    def _at_least_two_teams(cls, v: list[TeamInput]) -> list[TeamInput]:
        if len(v) < 2:
            raise ValueError("An event needs at least two teams")
        return v


class PatchTeamEventRequest(BaseModel):
    requesting_user_id: UUID
    title: str | None = None
    description: str | None = None
    status: str | None = None
    starts_at: datetime | None = None
    ends_at: datetime | None = None
    submission_mode: str | None = None
    requires_photo: bool | None = None
    image_url: str | None = None

    @field_validator("status")
    @classmethod
    def _valid_status(cls, v: str | None) -> str | None:
        if v is not None and v not in ("draft", "active", "completed", "cancelled"):
            raise ValueError("Invalid status")
        return v


class AddTeamRequest(BaseModel):
    requesting_user_id: UUID
    name: str
    color: str | None = None
    logo_url: str | None = None


class PatchTeamRequest(BaseModel):
    requesting_user_id: UUID
    name: str | None = None
    color: str | None = None
    logo_url: str | None = None


class JoinEventRequest(BaseModel):
    team_id: UUID | None = None
    participant_type: str
    user_id: UUID | None = None
    group_id: UUID | None = None
    # user joins only: opt in representing a group instead of as an individual.
    representing_group_id: UUID | None = None
    # group joins only: whether members are auto-enrolled or opt in themselves.
    cascade_mode: str = "cascade_all_members"
    requesting_user_id: UUID

    @model_validator(mode="after")
    def _one_of(self):
        if self.participant_type == "user":
            if not self.user_id or self.group_id:
                raise ValueError("participant_type 'user' requires user_id and no group_id")
            if not self.representing_group_id and not self.team_id:
                raise ValueError("team_id is required unless representing_group_id is set")
        elif self.participant_type == "group":
            if not self.group_id or self.user_id or self.representing_group_id:
                raise ValueError("participant_type 'group' requires group_id and no user_id/representing_group_id")
            if not self.team_id:
                raise ValueError("team_id is required")
            if self.cascade_mode not in ("cascade_all_members", "individual_opt_in"):
                raise ValueError("cascade_mode must be 'cascade_all_members' or 'individual_opt_in'")
        else:
            raise ValueError("participant_type must be 'user' or 'group'")
        return self


class OrganizerRequest(BaseModel):
    requesting_user_id: UUID
    user_id: UUID


class PatchSubmissionRequest(BaseModel):
    requesting_user_id: UUID
    small_bags: int | None = None
    large_bags: int | None = None
    pounds: float | None = None
    value: float | None = None
    review_status: str | None = None

    @field_validator("review_status")
    @classmethod
    def _valid_review_status(cls, v: str | None) -> str | None:
        if v is not None and v not in ("pending", "approved", "flagged"):
            raise ValueError("Invalid review_status")
        return v


@router.post("")
async def create_team_event(payload: CreateTeamEventRequest, db: AsyncSession = Depends(get_db)):
    if not await has_admin_role(db, payload.requesting_user_id, "event_manager"):
        raise HTTPException(status_code=403, detail="Only an event manager can create team events")

    existing = await db.execute(text("SELECT 1 FROM team_events WHERE slug = :slug"), {"slug": payload.slug})
    if existing.fetchone():
        raise HTTPException(status_code=409, detail="An event with this slug already exists")

    result = await db.execute(
        text("""
            INSERT INTO team_events
                (campaign_id, slug, title, description, starts_at, ends_at,
                 submission_mode, requires_photo, image_url, created_by)
            VALUES
                (:campaign_id, :slug, :title, :description, :starts_at, :ends_at,
                 :submission_mode, :requires_photo, :image_url, :created_by)
            RETURNING id
        """),
        {
            "campaign_id": str(payload.campaign_id) if payload.campaign_id else None,
            "slug": payload.slug,
            "title": payload.title,
            "description": payload.description,
            "starts_at": payload.starts_at,
            "ends_at": payload.ends_at,
            "submission_mode": payload.submission_mode,
            "requires_photo": payload.requires_photo,
            "image_url": payload.image_url,
            "created_by": str(payload.requesting_user_id),
        },
    )
    event_id = result.fetchone()[0]

    team_ids = []
    for team in payload.teams:
        team_result = await db.execute(
            text("""
                INSERT INTO team_event_teams (team_event_id, name, color, logo_url)
                VALUES (:team_event_id, :name, :color, :logo_url)
                RETURNING id
            """),
            {"team_event_id": str(event_id), "name": team.name, "color": team.color, "logo_url": team.logo_url},
        )
        team_ids.append(str(team_result.fetchone()[0]))

    await db.commit()
    return {"id": str(event_id), "slug": payload.slug, "team_ids": team_ids}


@router.get("")
async def list_team_events(requesting_user_id: UUID | None = None, db: AsyncSession = Depends(get_db)):
    is_manager = requesting_user_id is not None and await has_admin_role(db, requesting_user_id, "event_manager")
    if is_manager:
        result = await db.execute(text("SELECT id, slug, title, status, starts_at, ends_at, image_url FROM team_events ORDER BY starts_at DESC"))
    else:
        result = await db.execute(
            text("SELECT id, slug, title, status, starts_at, ends_at, image_url FROM team_events WHERE status != 'draft' ORDER BY starts_at DESC")
        )
    return [
        {
            "id": str(r.id),
            "slug": r.slug,
            "title": r.title,
            "status": r.status,
            "starts_at": r.starts_at,
            "ends_at": r.ends_at,
            "image_url": r.image_url,
        }
        for r in result.fetchall()
    ]


@router.get("/{team_event_id}")
async def get_team_event(team_event_id: UUID, db: AsyncSession = Depends(get_db)):
    event = await _get_event_or_404(db, team_event_id)
    teams_result = await db.execute(
        text("SELECT id, name, color, logo_url, geo_unit_id FROM team_event_teams WHERE team_event_id = :id ORDER BY name"),
        {"id": str(team_event_id)},
    )
    teams = [
        {"id": str(t.id), "name": t.name, "color": t.color, "logo_url": t.logo_url, "has_boundary": t.geo_unit_id is not None}
        for t in teams_result.fetchall()
    ]
    organizers_result = await db.execute(
        text("""
            SELECT teo.user_id, p.username, u.email
            FROM team_event_organizers teo
            JOIN profiles p ON p.id = teo.user_id
            JOIN auth.users u ON u.id = teo.user_id
            WHERE teo.team_event_id = :id
            ORDER BY teo.created_at
        """),
        {"id": str(team_event_id)},
    )
    organizers = [
        {"user_id": str(o.user_id), "username": o.username, "email": o.email}
        for o in organizers_result.fetchall()
    ]
    group_participants_result = await db.execute(
        text("""
            SELECT tegp.group_id, g.name AS group_name, tegp.team_id, tegp.cascade_mode
            FROM team_event_group_participants tegp
            JOIN groups g ON g.id = tegp.group_id
            WHERE tegp.team_event_id = :id
            ORDER BY g.name
        """),
        {"id": str(team_event_id)},
    )
    group_participants = [
        {
            "group_id": str(g.group_id),
            "group_name": g.group_name,
            "team_id": str(g.team_id),
            "cascade_mode": g.cascade_mode,
        }
        for g in group_participants_result.fetchall()
    ]
    return {
        "id": str(event.id),
        "campaign_id": str(event.campaign_id) if event.campaign_id else None,
        "slug": event.slug,
        "title": event.title,
        "description": event.description,
        "status": event.status,
        "starts_at": event.starts_at,
        "ends_at": event.ends_at,
        "submission_mode": event.submission_mode,
        "requires_photo": event.requires_photo,
        "image_url": event.image_url,
        "teams": teams,
        "organizers": organizers,
        "group_participants": group_participants,
    }


@router.get("/{team_event_id}/my-groups")
async def get_my_groups(team_event_id: UUID, user_id: UUID, db: AsyncSession = Depends(get_db)):
    await _get_event_or_404(db, team_event_id)
    result = await db.execute(
        text("""
            SELECT g.id, g.name, gm.role, tegp.team_id, tegp.cascade_mode
            FROM group_members gm
            JOIN groups g ON g.id = gm.group_id
            LEFT JOIN team_event_group_participants tegp
                ON tegp.group_id = g.id AND tegp.team_event_id = :event_id
            WHERE gm.user_id = :user_id
            ORDER BY g.name
        """),
        {"event_id": str(team_event_id), "user_id": str(user_id)},
    )
    return [
        {
            "group_id": str(r.id),
            "group_name": r.name,
            "is_admin": r.role == "admin",
            "joined_team_id": str(r.team_id) if r.team_id else None,
            "cascade_mode": r.cascade_mode,
        }
        for r in result.fetchall()
    ]


@router.get("/{team_event_id}/can-review")
async def get_can_review(team_event_id: UUID, user_id: UUID, db: AsyncSession = Depends(get_db)):
    await _get_event_or_404(db, team_event_id)
    return {"can_review": await _can_manage_team_event(db, team_event_id, user_id)}


@router.get("/{team_event_id}/stats")
async def get_team_event_stats(team_event_id: UUID, db: AsyncSession = Depends(get_db)):
    await _get_event_or_404(db, team_event_id)

    teams_result = await db.execute(
        text("""
            SELECT t.id AS team_id, t.name, t.color,
                   COALESCE(SUM(c.value) FILTER (WHERE c.review_status IS DISTINCT FROM 'flagged'), 0) AS total_value,
                   COUNT(c.id) FILTER (WHERE c.review_status IS DISTINCT FROM 'flagged') AS submission_count
            FROM team_event_teams t
            LEFT JOIN contributions c
                ON c.team_event_team_id = t.id AND c.team_event_id = t.team_event_id
            WHERE t.team_event_id = :id
            GROUP BY t.id, t.name, t.color
            ORDER BY total_value DESC
        """),
        {"id": str(team_event_id)},
    )
    teams = {
        str(r.team_id): {
            "team_id": str(r.team_id),
            "name": r.name,
            "color": r.color,
            "total_value": float(r.total_value),
            "submission_count": r.submission_count,
            "groups": [],
            "individuals": [],
        }
        for r in teams_result.fetchall()
    }

    roster_result = await db.execute(
        text("""
            SELECT tep.team_id, tep.representing_group_id AS group_id, g.name AS group_name, COUNT(*) AS member_count
            FROM team_event_participants tep
            JOIN groups g ON g.id = tep.representing_group_id
            WHERE tep.team_event_id = :id
            GROUP BY tep.team_id, tep.representing_group_id, g.name
        """),
        {"id": str(team_event_id)},
    )
    group_stats: dict[tuple[str, str], dict] = {}
    for r in roster_result.fetchall():
        group_stats[(str(r.team_id), str(r.group_id))] = {
            "group_id": str(r.group_id),
            "group_name": r.group_name,
            "member_count": r.member_count,
            "total_value": 0.0,
            "submission_count": 0,
        }

    group_rollup_result = await db.execute(
        text("""
            SELECT c.team_event_team_id AS team_id, tep.representing_group_id AS group_id,
                   COALESCE(SUM(c.value) FILTER (WHERE c.review_status IS DISTINCT FROM 'flagged'), 0) AS total_value,
                   COUNT(c.id) FILTER (WHERE c.review_status IS DISTINCT FROM 'flagged') AS submission_count
            FROM contributions c
            JOIN team_event_participants tep
                ON tep.team_event_id = c.team_event_id AND tep.user_id = c.user_id
            WHERE c.team_event_id = :id AND tep.representing_group_id IS NOT NULL
            GROUP BY c.team_event_team_id, tep.representing_group_id
        """),
        {"id": str(team_event_id)},
    )
    for r in group_rollup_result.fetchall():
        key = (str(r.team_id), str(r.group_id))
        if key in group_stats:
            group_stats[key]["total_value"] = float(r.total_value)
            group_stats[key]["submission_count"] = r.submission_count

    for (team_id, _group_id), g in group_stats.items():
        if team_id in teams:
            teams[team_id]["groups"].append(g)

    individuals_result = await db.execute(
        text("""
            SELECT tep.team_id, tep.user_id, p.username,
                   COALESCE(SUM(c.value) FILTER (WHERE c.review_status IS DISTINCT FROM 'flagged'), 0) AS total_value,
                   COUNT(c.id) FILTER (WHERE c.review_status IS DISTINCT FROM 'flagged') AS submission_count
            FROM team_event_participants tep
            JOIN profiles p ON p.id = tep.user_id
            LEFT JOIN contributions c
                ON c.team_event_id = tep.team_event_id AND c.user_id = tep.user_id
            WHERE tep.team_event_id = :id AND tep.representing_group_id IS NULL
            GROUP BY tep.team_id, tep.user_id, p.username
        """),
        {"id": str(team_event_id)},
    )
    for r in individuals_result.fetchall():
        team_id = str(r.team_id)
        if team_id in teams:
            teams[team_id]["individuals"].append({
                "user_id": str(r.user_id),
                "username": r.username,
                "total_value": float(r.total_value),
                "submission_count": r.submission_count,
            })

    for t in teams.values():
        t["groups"].sort(key=lambda g: g["total_value"], reverse=True)
        t["individuals"].sort(key=lambda i: i["total_value"], reverse=True)

    return sorted(teams.values(), key=lambda t: t["total_value"], reverse=True)


@router.get("/{team_event_id}/leaderboard")
async def get_team_event_leaderboard(
    team_event_id: UUID,
    scope: str = Query("individuals", pattern="^(individuals|groups)$"),
    interval: str = Query("all"),
    start_date: str | None = None,
    end_date: str | None = None,
    db: AsyncSession = Depends(get_db),
):
    await _get_event_or_404(db, team_event_id)
    start, end = await resolve_stats_window(db, interval, start_date, end_date)
    params = {"id": str(team_event_id), "start": start, "end": end}

    if scope == "groups":
        result = await db.execute(
            text("""
                SELECT tegp.group_id, g.name AS group_name, g.slug, g.image_url AS logo_url,
                       tegp.team_id, t.name AS team_name, t.color AS team_color,
                       COUNT(DISTINCT tep.user_id) AS member_count,
                       COALESCE(SUM(c.value) FILTER (
                           WHERE c.review_status IS DISTINCT FROM 'flagged'
                           AND (CAST(:start AS timestamptz) IS NULL OR c.submitted_at >= :start)
                           AND (CAST(:end AS timestamptz) IS NULL OR c.submitted_at < :end)
                       ), 0) AS total_value,
                       COUNT(c.id) FILTER (
                           WHERE c.review_status IS DISTINCT FROM 'flagged'
                           AND (CAST(:start AS timestamptz) IS NULL OR c.submitted_at >= :start)
                           AND (CAST(:end AS timestamptz) IS NULL OR c.submitted_at < :end)
                       ) AS submission_count
                FROM team_event_group_participants tegp
                JOIN groups g ON g.id = tegp.group_id
                JOIN team_event_teams t ON t.id = tegp.team_id
                LEFT JOIN team_event_participants tep
                    ON tep.team_event_id = tegp.team_event_id AND tep.representing_group_id = tegp.group_id
                LEFT JOIN contributions c
                    ON c.team_event_id = tegp.team_event_id AND c.user_id = tep.user_id
                WHERE tegp.team_event_id = :id
                GROUP BY tegp.group_id, g.name, g.slug, g.image_url, tegp.team_id, t.name, t.color
                ORDER BY total_value DESC
            """),
            params,
        )
        return [
            {
                "group_id": str(r.group_id),
                "group_name": r.group_name,
                "slug": r.slug,
                "logo_url": r.logo_url,
                "team_id": str(r.team_id),
                "team_name": r.team_name,
                "team_color": r.team_color,
                "member_count": r.member_count,
                "total_value": float(r.total_value),
                "submission_count": r.submission_count,
            }
            for r in result.fetchall()
        ]

    result = await db.execute(
        text("""
            SELECT tep.user_id, p.username, p.display_name, p.avatar_url,
                   tep.team_id, t.name AS team_name, t.color AS team_color,
                   COALESCE(SUM(c.value) FILTER (
                       WHERE c.review_status IS DISTINCT FROM 'flagged'
                       AND (CAST(:start AS timestamptz) IS NULL OR c.submitted_at >= :start)
                       AND (CAST(:end AS timestamptz) IS NULL OR c.submitted_at < :end)
                   ), 0) AS total_value,
                   COUNT(c.id) FILTER (
                       WHERE c.review_status IS DISTINCT FROM 'flagged'
                       AND (CAST(:start AS timestamptz) IS NULL OR c.submitted_at >= :start)
                       AND (CAST(:end AS timestamptz) IS NULL OR c.submitted_at < :end)
                   ) AS submission_count
            FROM team_event_participants tep
            JOIN profiles p ON p.id = tep.user_id
            JOIN team_event_teams t ON t.id = tep.team_id
            LEFT JOIN contributions c
                ON c.team_event_id = tep.team_event_id AND c.user_id = tep.user_id
            WHERE tep.team_event_id = :id AND tep.representing_group_id IS NULL
            GROUP BY tep.user_id, p.username, p.display_name, p.avatar_url, tep.team_id, t.name, t.color
            ORDER BY total_value DESC
        """),
        params,
    )
    return [
        {
            "user_id": str(r.user_id),
            "username": r.username,
            "display_name": r.display_name,
            "avatar_url": r.avatar_url,
            "team_id": str(r.team_id),
            "team_name": r.team_name,
            "team_color": r.team_color,
            "total_value": float(r.total_value),
            "submission_count": r.submission_count,
        }
        for r in result.fetchall()
    ]


@router.get("/{team_event_id}/geo")
async def get_team_event_geo(team_event_id: UUID, db: AsyncSession = Depends(get_db)):
    await _get_event_or_404(db, team_event_id)
    result = await db.execute(
        text("""
            SELECT t.id AS team_id, t.name AS team_name, t.color AS team_color,
                   gu.id AS geo_unit_id, gu.display_name AS geo_display_name, gu.unit_type,
                   COALESCE(SUM(c.value) FILTER (WHERE c.review_status IS DISTINCT FROM 'flagged'), 0) AS total_value,
                   COUNT(c.id) FILTER (WHERE c.review_status IS DISTINCT FROM 'flagged') AS submission_count
            FROM team_event_teams t
            JOIN geo_units gu ON gu.id = t.geo_unit_id
            LEFT JOIN contributions c
                ON c.team_event_team_id = t.id AND c.team_event_id = t.team_event_id
            WHERE t.team_event_id = :id AND t.geo_unit_id IS NOT NULL
            GROUP BY t.id, t.name, t.color, gu.id, gu.display_name, gu.unit_type
            ORDER BY total_value DESC
        """),
        {"id": str(team_event_id)},
    )
    return [
        {
            "team_id": str(r.team_id),
            "team_name": r.team_name,
            "team_color": r.team_color,
            "geo_unit_id": str(r.geo_unit_id),
            "geo_display_name": r.geo_display_name,
            "unit_type": r.unit_type,
            "total_value": float(r.total_value),
            "submission_count": r.submission_count,
        }
        for r in result.fetchall()
    ]


@router.get("/{team_event_id}/participant-detail")
async def get_team_event_participant_detail(
    team_event_id: UUID,
    type: str = Query(..., pattern="^(user|group)$"),
    id: UUID = Query(...),
    interval: str = Query("all"),
    start_date: str | None = None,
    end_date: str | None = None,
    db: AsyncSession = Depends(get_db),
):
    await _get_event_or_404(db, team_event_id)
    start, end = await resolve_stats_window(db, interval, start_date, end_date)
    bucket = trend_bucket_unit(start, end)
    params = {"id": str(team_event_id), "entity_id": str(id), "start": start, "end": end, "bucket": bucket}

    if type == "group":
        identity_result = await db.execute(
            text("SELECT name, slug, image_url AS logo_url FROM groups WHERE id = :entity_id"),
            {"entity_id": str(id)},
        )
        identity_row = identity_result.fetchone()
        if not identity_row:
            raise HTTPException(status_code=404, detail="Group not found")
        identity = {"name": identity_row.name, "slug": identity_row.slug, "logo_url": identity_row.logo_url}
        contributor_filter = """
            EXISTS (
                SELECT 1 FROM team_event_participants tep
                WHERE tep.team_event_id = c.team_event_id AND tep.user_id = c.user_id
                AND tep.representing_group_id = :entity_id
            )
        """
    else:
        identity_result = await db.execute(
            text("SELECT username, display_name, avatar_url FROM profiles WHERE id = :entity_id"),
            {"entity_id": str(id)},
        )
        identity_row = identity_result.fetchone()
        if not identity_row:
            raise HTTPException(status_code=404, detail="User not found")
        identity = {
            "username": identity_row.username,
            "display_name": identity_row.display_name,
            "avatar_url": identity_row.avatar_url,
        }
        contributor_filter = "c.user_id = :entity_id"

    summary_result = await db.execute(
        text(f"""
            SELECT COALESCE(SUM(c.value) FILTER (WHERE c.review_status IS DISTINCT FROM 'flagged'), 0) AS total_value,
                   COUNT(c.id) FILTER (WHERE c.review_status IS DISTINCT FROM 'flagged') AS submission_count
            FROM contributions c
            WHERE c.team_event_id = :id AND {contributor_filter}
              AND (CAST(:start AS timestamptz) IS NULL OR c.submitted_at >= :start)
              AND (CAST(:end AS timestamptz) IS NULL OR c.submitted_at < :end)
        """),
        params,
    )
    summary_row = summary_result.fetchone()

    breakdown_result = await db.execute(
        text(f"""
            SELECT c.contribution_type,
                   COALESCE(SUM(c.value), 0) AS total_value,
                   COUNT(c.id) AS submission_count
            FROM contributions c
            WHERE c.team_event_id = :id AND {contributor_filter}
              AND c.review_status IS DISTINCT FROM 'flagged'
              AND (CAST(:start AS timestamptz) IS NULL OR c.submitted_at >= :start)
              AND (CAST(:end AS timestamptz) IS NULL OR c.submitted_at < :end)
            GROUP BY c.contribution_type
            ORDER BY total_value DESC
        """),
        params,
    )

    trend_result = await db.execute(
        text(f"""
            SELECT date_trunc(:bucket, c.submitted_at) AS bucket_start,
                   COALESCE(SUM(c.value), 0) AS total_value
            FROM contributions c
            WHERE c.team_event_id = :id AND {contributor_filter}
              AND c.review_status IS DISTINCT FROM 'flagged'
              AND (CAST(:start AS timestamptz) IS NULL OR c.submitted_at >= :start)
              AND (CAST(:end AS timestamptz) IS NULL OR c.submitted_at < :end)
            GROUP BY bucket_start
            ORDER BY bucket_start
        """),
        params,
    )

    return {
        "type": type,
        "id": str(id),
        "identity": identity,
        "total_value": float(summary_row.total_value),
        "submission_count": summary_row.submission_count,
        "breakdown": [
            {"contribution_type": r.contribution_type, "total_value": float(r.total_value), "submission_count": r.submission_count}
            for r in breakdown_result.fetchall()
        ],
        "trend": [
            {"bucket_start": r.bucket_start, "total_value": float(r.total_value)}
            for r in trend_result.fetchall()
        ],
    }


@router.get("/{team_event_id}/admin-summary")
async def get_team_event_admin_summary(
    team_event_id: UUID,
    interval: str = Query("all"),
    start_date: str | None = None,
    end_date: str | None = None,
    db: AsyncSession = Depends(get_db),
):
    await _get_event_or_404(db, team_event_id)

    start, end = await resolve_stats_window(db, interval, start_date, end_date)
    bucket = trend_bucket_unit(start, end)
    params = {"id": str(team_event_id), "start": start, "end": end, "bucket": bucket}

    kpi_result = await db.execute(
        text("""
            SELECT COALESCE(SUM(c.value) FILTER (WHERE c.review_status IS DISTINCT FROM 'flagged'), 0) AS total_value,
                   COUNT(c.id) FILTER (WHERE c.review_status IS DISTINCT FROM 'flagged') AS submission_count,
                   COUNT(DISTINCT c.user_id) FILTER (WHERE c.review_status IS DISTINCT FROM 'flagged') AS active_participants,
                   COUNT(c.id) FILTER (WHERE c.review_status = 'pending') AS pending_review_count
            FROM contributions c
            WHERE c.team_event_id = :id
              AND (CAST(:start AS timestamptz) IS NULL OR c.submitted_at >= :start)
              AND (CAST(:end AS timestamptz) IS NULL OR c.submitted_at < :end)
        """),
        params,
    )
    kpi_row = kpi_result.fetchone()

    roster_result = await db.execute(
        text("""
            SELECT COUNT(DISTINCT tep.user_id) AS total_participants,
                   COUNT(DISTINCT tep.representing_group_id) FILTER (WHERE tep.representing_group_id IS NOT NULL) AS total_groups,
                   COUNT(DISTINCT tep.team_id) AS total_teams
            FROM team_event_participants tep
            WHERE tep.team_event_id = :id
        """),
        {"id": str(team_event_id)},
    )
    roster_row = roster_result.fetchone()

    trend_result = await db.execute(
        text("""
            SELECT date_trunc(:bucket, c.submitted_at) AS bucket_start,
                   COALESCE(SUM(c.value), 0) AS total_value,
                   COUNT(c.id) AS submission_count
            FROM contributions c
            WHERE c.team_event_id = :id AND c.review_status IS DISTINCT FROM 'flagged'
              AND (CAST(:start AS timestamptz) IS NULL OR c.submitted_at >= :start)
              AND (CAST(:end AS timestamptz) IS NULL OR c.submitted_at < :end)
            GROUP BY bucket_start
            ORDER BY bucket_start
        """),
        params,
    )

    breakdown_result = await db.execute(
        text("""
            SELECT c.contribution_type, COALESCE(SUM(c.value), 0) AS total_value, COUNT(c.id) AS submission_count
            FROM contributions c
            WHERE c.team_event_id = :id AND c.review_status IS DISTINCT FROM 'flagged'
              AND (CAST(:start AS timestamptz) IS NULL OR c.submitted_at >= :start)
              AND (CAST(:end AS timestamptz) IS NULL OR c.submitted_at < :end)
            GROUP BY c.contribution_type
            ORDER BY total_value DESC
        """),
        params,
    )

    team_totals_result = await db.execute(
        text("""
            SELECT t.id AS team_id, t.name, t.color,
                   COALESCE(SUM(c.value) FILTER (
                       WHERE c.review_status IS DISTINCT FROM 'flagged'
                       AND (CAST(:start AS timestamptz) IS NULL OR c.submitted_at >= :start)
                       AND (CAST(:end AS timestamptz) IS NULL OR c.submitted_at < :end)
                   ), 0) AS total_value,
                   COUNT(c.id) FILTER (
                       WHERE c.review_status IS DISTINCT FROM 'flagged'
                       AND (CAST(:start AS timestamptz) IS NULL OR c.submitted_at >= :start)
                       AND (CAST(:end AS timestamptz) IS NULL OR c.submitted_at < :end)
                   ) AS submission_count
            FROM team_event_teams t
            LEFT JOIN contributions c ON c.team_event_team_id = t.id AND c.team_event_id = t.team_event_id
            WHERE t.team_event_id = :id
            GROUP BY t.id, t.name, t.color
        """),
        params,
    )
    team_totals = {
        str(r.team_id): {
            "team_id": str(r.team_id),
            "name": r.name,
            "color": r.color,
            "total_value": float(r.total_value),
            "submission_count": r.submission_count,
            "participant_count": 0,
        }
        for r in team_totals_result.fetchall()
    }

    team_roster_result = await db.execute(
        text("""
            SELECT tep.team_id, COUNT(DISTINCT tep.user_id) AS participant_count
            FROM team_event_participants tep
            WHERE tep.team_event_id = :id
            GROUP BY tep.team_id
        """),
        {"id": str(team_event_id)},
    )
    for r in team_roster_result.fetchall():
        team_id = str(r.team_id)
        if team_id in team_totals:
            team_totals[team_id]["participant_count"] = r.participant_count

    teams_comparison = sorted(team_totals.values(), key=lambda t: t["total_value"], reverse=True)

    top_contributors_result = await db.execute(
        text("""
            SELECT c.user_id, p.username, p.display_name, p.avatar_url,
                   COALESCE(SUM(c.value), 0) AS total_value, COUNT(c.id) AS submission_count
            FROM contributions c
            JOIN profiles p ON p.id = c.user_id
            WHERE c.team_event_id = :id AND c.review_status IS DISTINCT FROM 'flagged'
              AND (CAST(:start AS timestamptz) IS NULL OR c.submitted_at >= :start)
              AND (CAST(:end AS timestamptz) IS NULL OR c.submitted_at < :end)
            GROUP BY c.user_id, p.username, p.display_name, p.avatar_url
            ORDER BY total_value DESC
            LIMIT 10
        """),
        params,
    )

    return {
        "total_value": float(kpi_row.total_value),
        "submission_count": kpi_row.submission_count,
        "active_participants": kpi_row.active_participants,
        "pending_review_count": kpi_row.pending_review_count,
        "total_participants": roster_row.total_participants,
        "total_groups": roster_row.total_groups,
        "total_teams": roster_row.total_teams,
        "trend": [
            {"bucket_start": r.bucket_start, "total_value": float(r.total_value), "submission_count": r.submission_count}
            for r in trend_result.fetchall()
        ],
        "breakdown": [
            {"contribution_type": r.contribution_type, "total_value": float(r.total_value), "submission_count": r.submission_count}
            for r in breakdown_result.fetchall()
        ],
        "teams": teams_comparison,
        "top_contributors": [
            {
                "user_id": str(r.user_id),
                "username": r.username,
                "display_name": r.display_name,
                "avatar_url": r.avatar_url,
                "total_value": float(r.total_value),
                "submission_count": r.submission_count,
            }
            for r in top_contributors_result.fetchall()
        ],
    }


@router.get("/{team_event_id}/export.csv")
async def export_team_event_summary_csv(
    team_event_id: UUID,
    interval: str = Query("all"),
    start_date: str | None = None,
    end_date: str | None = None,
    db: AsyncSession = Depends(get_db),
):
    """CSV export of the public stats summary -- teams, breakdown, and top contributors."""
    summary = await get_team_event_admin_summary(team_event_id, interval, start_date, end_date, db)

    buf = io.StringIO()
    writer = csv.writer(buf)

    writer.writerow(["Teams"])
    writer.writerow(["team", "total_value", "submission_count", "participant_count"])
    for t in summary["teams"]:
        writer.writerow([t["name"], t["total_value"], t["submission_count"], t["participant_count"]])

    writer.writerow([])
    writer.writerow(["Breakdown by type"])
    writer.writerow(["contribution_type", "total_value", "submission_count"])
    for b in summary["breakdown"]:
        writer.writerow([b["contribution_type"], b["total_value"], b["submission_count"]])

    writer.writerow([])
    writer.writerow(["Top contributors"])
    writer.writerow(["username", "display_name", "total_value", "submission_count"])
    for c in summary["top_contributors"]:
        writer.writerow([c["username"], c["display_name"], c["total_value"], c["submission_count"]])

    filename = f"team-event-{team_event_id}-stats-{interval}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{team_event_id}/export.xlsx")
async def export_team_event_summary_xlsx(
    team_event_id: UUID,
    interval: str = Query("all"),
    start_date: str | None = None,
    end_date: str | None = None,
    db: AsyncSession = Depends(get_db),
):
    """Excel export of the public stats summary -- same shape as the CSV export."""
    summary = await get_team_event_admin_summary(team_event_id, interval, start_date, end_date, db)

    wb = Workbook()

    ws_teams = wb.active
    ws_teams.title = "Teams"
    ws_teams.append(["team", "total_value", "submission_count", "participant_count"])
    for t in summary["teams"]:
        ws_teams.append([t["name"], t["total_value"], t["submission_count"], t["participant_count"]])

    ws_breakdown = wb.create_sheet("Breakdown")
    ws_breakdown.append(["contribution_type", "total_value", "submission_count"])
    for b in summary["breakdown"]:
        ws_breakdown.append([b["contribution_type"], b["total_value"], b["submission_count"]])

    ws_contributors = wb.create_sheet("Top Contributors")
    ws_contributors.append(["username", "display_name", "total_value", "submission_count"])
    for c in summary["top_contributors"]:
        ws_contributors.append([c["username"], c["display_name"], c["total_value"], c["submission_count"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"team-event-{team_event_id}-stats-{interval}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


async def _fetch_team_event_full_export_data(db: AsyncSession, team_event_id: UUID) -> dict:
    teams_result = await db.execute(
        text("SELECT id, name, color, logo_url FROM team_event_teams WHERE team_event_id = :id ORDER BY name"),
        {"id": str(team_event_id)},
    )
    teams = {str(r.id): {"team_id": str(r.id), "name": r.name, "color": r.color, "logo_url": r.logo_url} for r in teams_result.fetchall()}

    participants_result = await db.execute(
        text("""
            SELECT tep.user_id, p.username, p.display_name, tep.team_id, t.name AS team_name,
                   tep.representing_group_id, g.name AS group_name, tep.joined_at
            FROM team_event_participants tep
            JOIN profiles p ON p.id = tep.user_id
            JOIN team_event_teams t ON t.id = tep.team_id
            LEFT JOIN groups g ON g.id = tep.representing_group_id
            WHERE tep.team_event_id = :id
            ORDER BY t.name, p.username
        """),
        {"id": str(team_event_id)},
    )
    participants = [
        {
            "user_id": str(r.user_id),
            "username": r.username,
            "display_name": r.display_name,
            "team_name": r.team_name,
            "representing_group": r.group_name,
            "joined_at": r.joined_at.isoformat() if r.joined_at else None,
        }
        for r in participants_result.fetchall()
    ]

    submissions_result = await db.execute(
        text("""
            SELECT c.id, c.submitted_at, c.user_id, p.username, p.display_name,
                   c.team_event_team_id, t.name AS team_name, c.contribution_type, c.value, c.review_status
            FROM contributions c
            JOIN profiles p ON p.id = c.user_id
            LEFT JOIN team_event_teams t ON t.id = c.team_event_team_id
            WHERE c.team_event_id = :id
            ORDER BY c.submitted_at DESC
        """),
        {"id": str(team_event_id)},
    )
    submissions = [
        {
            "id": str(r.id),
            "submitted_at": r.submitted_at.isoformat() if r.submitted_at else None,
            "username": r.username,
            "display_name": r.display_name,
            "team_name": r.team_name,
            "contribution_type": r.contribution_type,
            "value": float(r.value) if r.value is not None else 0.0,
            "review_status": r.review_status,
        }
        for r in submissions_result.fetchall()
    ]

    return {"teams": list(teams.values()), "participants": participants, "submissions": submissions}


@router.get("/{team_event_id}/export/full.csv")
async def export_team_event_full_csv(
    team_event_id: UUID,
    requesting_user_id: UUID = Query(...),
    db: AsyncSession = Depends(get_db),
):
    """Full raw-data CSV export (all participants + all submissions) -- event organizers only."""
    await _get_event_or_404(db, team_event_id)
    if not await _can_manage_team_event(db, team_event_id, requesting_user_id):
        raise HTTPException(status_code=403, detail="Only event organizers can export the full event data.")

    data = await _fetch_team_event_full_export_data(db, team_event_id)

    buf = io.StringIO()
    writer = csv.writer(buf)

    writer.writerow(["Teams"])
    writer.writerow(["team", "color"])
    for t in data["teams"]:
        writer.writerow([t["name"], t["color"]])

    writer.writerow([])
    writer.writerow(["Participants"])
    writer.writerow(["username", "display_name", "team", "representing_group", "joined_at"])
    for p in data["participants"]:
        writer.writerow([p["username"], p["display_name"], p["team_name"], p["representing_group"], p["joined_at"]])

    writer.writerow([])
    writer.writerow(["Submissions"])
    writer.writerow(["submitted_at", "username", "display_name", "team", "contribution_type", "value", "review_status"])
    for s in data["submissions"]:
        writer.writerow([s["submitted_at"], s["username"], s["display_name"], s["team_name"], s["contribution_type"], s["value"], s["review_status"]])

    filename = f"team-event-{team_event_id}-full-export.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{team_event_id}/export/full.xlsx")
async def export_team_event_full_xlsx(
    team_event_id: UUID,
    requesting_user_id: UUID = Query(...),
    db: AsyncSession = Depends(get_db),
):
    """Full raw-data Excel export (all participants + all submissions) -- event organizers only."""
    await _get_event_or_404(db, team_event_id)
    if not await _can_manage_team_event(db, team_event_id, requesting_user_id):
        raise HTTPException(status_code=403, detail="Only event organizers can export the full event data.")

    data = await _fetch_team_event_full_export_data(db, team_event_id)

    wb = Workbook()

    ws_teams = wb.active
    ws_teams.title = "Teams"
    ws_teams.append(["team", "color"])
    for t in data["teams"]:
        ws_teams.append([t["name"], t["color"]])

    ws_participants = wb.create_sheet("Participants")
    ws_participants.append(["username", "display_name", "team", "representing_group", "joined_at"])
    for p in data["participants"]:
        ws_participants.append([p["username"], p["display_name"], p["team_name"], p["representing_group"], p["joined_at"]])

    ws_submissions = wb.create_sheet("Submissions")
    ws_submissions.append(["submitted_at", "username", "display_name", "team", "contribution_type", "value", "review_status"])
    for s in data["submissions"]:
        ws_submissions.append([s["submitted_at"], s["username"], s["display_name"], s["team_name"], s["contribution_type"], s["value"], s["review_status"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"team-event-{team_event_id}-full-export.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.patch("/{team_event_id}")
async def patch_team_event(team_event_id: UUID, payload: PatchTeamEventRequest, db: AsyncSession = Depends(get_db)):
    await _get_event_or_404(db, team_event_id)
    if not await _can_manage_team_event(db, team_event_id, payload.requesting_user_id):
        raise HTTPException(status_code=403, detail="Not authorized to manage this event")

    fields = payload.model_dump(exclude={"requesting_user_id"}, exclude_unset=True)
    if not fields:
        return {"updated": False}

    set_clause = ", ".join(f"{k} = :{k}" for k in fields)
    await db.execute(
        text(f"UPDATE team_events SET {set_clause}, updated_at = now() WHERE id = :id"),
        {**fields, "id": str(team_event_id)},
    )
    await db.commit()
    return {"updated": True}


@router.post("/{team_event_id}/teams")
async def add_team(team_event_id: UUID, payload: AddTeamRequest, db: AsyncSession = Depends(get_db)):
    await _get_event_or_404(db, team_event_id)
    if not await _can_manage_team_event(db, team_event_id, payload.requesting_user_id):
        raise HTTPException(status_code=403, detail="Not authorized to manage this event")

    result = await db.execute(
        text("""
            INSERT INTO team_event_teams (team_event_id, name, color, logo_url)
            VALUES (:team_event_id, :name, :color, :logo_url)
            RETURNING id
        """),
        {"team_event_id": str(team_event_id), "name": payload.name, "color": payload.color, "logo_url": payload.logo_url},
    )
    team_id = result.fetchone()[0]
    await db.commit()
    return {"id": str(team_id)}


@router.patch("/{team_event_id}/teams/{team_id}")
async def patch_team(
    team_event_id: UUID, team_id: UUID, payload: PatchTeamRequest, db: AsyncSession = Depends(get_db)
):
    await _get_event_or_404(db, team_event_id)
    if not await _can_manage_team_event(db, team_event_id, payload.requesting_user_id):
        raise HTTPException(status_code=403, detail="Not authorized to manage this event")

    team_row = await db.execute(
        text("SELECT 1 FROM team_event_teams WHERE id = :id AND team_event_id = :event_id"),
        {"id": str(team_id), "event_id": str(team_event_id)},
    )
    if not team_row.fetchone():
        raise HTTPException(status_code=404, detail="Team not found for this event")

    fields = payload.model_dump(exclude={"requesting_user_id"}, exclude_unset=True)
    if not fields:
        return {"updated": False}

    set_clause = ", ".join(f"{k} = :{k}" for k in fields)
    await db.execute(
        text(f"UPDATE team_event_teams SET {set_clause} WHERE id = :id"),
        {**fields, "id": str(team_id)},
    )
    await db.commit()
    return {"updated": True}


@router.post("/{team_event_id}/join")
async def join_event(team_event_id: UUID, payload: JoinEventRequest, db: AsyncSession = Depends(get_db)):
    event = await _get_event_or_404(db, team_event_id)
    if event.status not in ("draft", "active"):
        raise HTTPException(status_code=400, detail="This event is no longer accepting participants")

    if payload.participant_type == "group":
        team_row = await db.execute(
            text("SELECT 1 FROM team_event_teams WHERE id = :team_id AND team_event_id = :event_id"),
            {"team_id": str(payload.team_id), "event_id": str(team_event_id)},
        )
        if not team_row.fetchone():
            raise HTTPException(status_code=404, detail="Team not found for this event")
        if not await _is_group_admin(db, payload.group_id, payload.requesting_user_id):
            raise HTTPException(status_code=403, detail="Only a group admin can opt the group in")

        await db.execute(
            text("""
                INSERT INTO team_event_group_participants (team_event_id, team_id, group_id, cascade_mode)
                VALUES (:team_event_id, :team_id, :group_id, :cascade_mode)
                ON CONFLICT (team_event_id, group_id)
                DO UPDATE SET team_id = EXCLUDED.team_id, cascade_mode = EXCLUDED.cascade_mode
            """),
            {
                "team_event_id": str(team_event_id),
                "team_id": str(payload.team_id),
                "group_id": str(payload.group_id),
                "cascade_mode": payload.cascade_mode,
            },
        )
        if payload.cascade_mode == "cascade_all_members":
            # Enroll current members as individual participants representing this
            # group, and re-team anyone already representing it if the group switched
            # teams. Never touches a member who opted in independently or is
            # representing a different group.
            await db.execute(
                text("""
                    INSERT INTO team_event_participants (team_event_id, team_id, user_id, representing_group_id)
                    SELECT :team_event_id, :team_id, gm.user_id, :group_id
                    FROM group_members gm WHERE gm.group_id = :group_id
                    ON CONFLICT (team_event_id, user_id) DO UPDATE
                    SET team_id = EXCLUDED.team_id
                    WHERE team_event_participants.representing_group_id = EXCLUDED.representing_group_id
                """),
                {"team_event_id": str(team_event_id), "team_id": str(payload.team_id), "group_id": str(payload.group_id)},
            )
        await db.commit()
        return {"joined": True, "team_id": str(payload.team_id)}

    if payload.user_id != payload.requesting_user_id:
        raise HTTPException(status_code=403, detail="Users can only opt themselves in")

    team_id = payload.team_id
    if payload.representing_group_id:
        member_row = await db.execute(
            text("SELECT 1 FROM group_members WHERE group_id = :group_id AND user_id = :user_id"),
            {"group_id": str(payload.representing_group_id), "user_id": str(payload.user_id)},
        )
        if not member_row.fetchone():
            raise HTTPException(status_code=403, detail="You're not a member of that group")
        group_join = await db.execute(
            text("SELECT team_id FROM team_event_group_participants WHERE team_event_id = :event_id AND group_id = :group_id"),
            {"event_id": str(team_event_id), "group_id": str(payload.representing_group_id)},
        )
        group_join_row = group_join.fetchone()
        if not group_join_row:
            raise HTTPException(status_code=400, detail="That group hasn't joined this event yet")
        team_id = group_join_row.team_id
    else:
        team_row = await db.execute(
            text("SELECT 1 FROM team_event_teams WHERE id = :team_id AND team_event_id = :event_id"),
            {"team_id": str(team_id), "event_id": str(team_event_id)},
        )
        if not team_row.fetchone():
            raise HTTPException(status_code=404, detail="Team not found for this event")

    await db.execute(
        text("""
            INSERT INTO team_event_participants (team_event_id, team_id, user_id, representing_group_id)
            VALUES (:team_event_id, :team_id, :user_id, :representing_group_id)
            ON CONFLICT (team_event_id, user_id)
            DO UPDATE SET team_id = EXCLUDED.team_id, representing_group_id = EXCLUDED.representing_group_id
        """),
        {
            "team_event_id": str(team_event_id),
            "team_id": str(team_id),
            "user_id": str(payload.user_id),
            "representing_group_id": str(payload.representing_group_id) if payload.representing_group_id else None,
        },
    )
    await db.commit()
    return {"joined": True, "team_id": str(team_id)}


@router.delete("/{team_event_id}/leave")
async def leave_event(
    team_event_id: UUID,
    requesting_user_id: UUID,
    group_id: UUID | None = None,
    db: AsyncSession = Depends(get_db),
):
    await _get_event_or_404(db, team_event_id)

    if group_id is not None:
        if not await _is_group_admin(db, group_id, requesting_user_id):
            raise HTTPException(status_code=403, detail="Only a group admin can withdraw the group")
        # Members auto-enrolled by this group's cascade leave with it; their past
        # contributions keep counting toward the team (attributed on the contribution
        # row itself), they just drop off the live roster and stop auto-attributing.
        await db.execute(
            text("DELETE FROM team_event_participants WHERE team_event_id = :event_id AND representing_group_id = :group_id"),
            {"event_id": str(team_event_id), "group_id": str(group_id)},
        )
        result = await db.execute(
            text("DELETE FROM team_event_group_participants WHERE team_event_id = :event_id AND group_id = :group_id"),
            {"event_id": str(team_event_id), "group_id": str(group_id)},
        )
    else:
        result = await db.execute(
            text("DELETE FROM team_event_participants WHERE team_event_id = :event_id AND user_id = :user_id"),
            {"event_id": str(team_event_id), "user_id": str(requesting_user_id)},
        )
    await db.commit()
    return {"left": result.rowcount > 0}


@router.post("/{team_event_id}/organizers")
async def add_organizer(team_event_id: UUID, payload: OrganizerRequest, db: AsyncSession = Depends(get_db)):
    await _get_event_or_404(db, team_event_id)
    if not await has_admin_role(db, payload.requesting_user_id, "event_manager"):
        raise HTTPException(status_code=403, detail="Only an event manager can delegate organizers")

    await db.execute(
        text("""
            INSERT INTO team_event_organizers (team_event_id, user_id, granted_by)
            VALUES (:team_event_id, :user_id, :granted_by)
            ON CONFLICT (team_event_id, user_id) DO NOTHING
        """),
        {"team_event_id": str(team_event_id), "user_id": str(payload.user_id), "granted_by": str(payload.requesting_user_id)},
    )
    await db.commit()
    return {"added": True}


@router.delete("/{team_event_id}/organizers/{user_id}")
async def remove_organizer(team_event_id: UUID, user_id: UUID, requesting_user_id: UUID, db: AsyncSession = Depends(get_db)):
    await _get_event_or_404(db, team_event_id)
    if not await has_admin_role(db, requesting_user_id, "event_manager"):
        raise HTTPException(status_code=403, detail="Only an event manager can remove organizers")
    await db.execute(
        text("DELETE FROM team_event_organizers WHERE team_event_id = :team_event_id AND user_id = :user_id"),
        {"team_event_id": str(team_event_id), "user_id": str(user_id)},
    )
    await db.commit()
    return {"removed": True}


@router.get("/{team_event_id}/scoreboard")
async def get_scoreboard(team_event_id: UUID, db: AsyncSession = Depends(get_db)):
    await _get_event_or_404(db, team_event_id)
    result = await db.execute(
        text("""
            SELECT t.id AS team_id, t.name, t.color,
                   COALESCE(SUM(c.value), 0) AS total_value,
                   COUNT(c.id) AS submission_count
            FROM team_event_teams t
            LEFT JOIN contributions c
                ON c.team_event_team_id = t.id
                AND c.team_event_id = t.team_event_id
                AND (c.review_status IS DISTINCT FROM 'flagged')
            WHERE t.team_event_id = :id
            GROUP BY t.id, t.name, t.color
            ORDER BY total_value DESC
        """),
        {"id": str(team_event_id)},
    )
    return [
        {"team_id": str(r.team_id), "name": r.name, "color": r.color, "total_value": float(r.total_value), "submission_count": r.submission_count}
        for r in result.fetchall()
    ]


@router.get("/{team_event_id}/submissions")
async def list_submissions(team_event_id: UUID, requesting_user_id: UUID, db: AsyncSession = Depends(get_db)):
    await _get_event_or_404(db, team_event_id)
    if not await _can_manage_team_event(db, team_event_id, requesting_user_id):
        raise HTTPException(status_code=403, detail="Not authorized to review this event's submissions")

    result = await db.execute(
        text("""
            SELECT c.id, c.user_id, c.team_event_team_id, c.contribution_type, c.value,
                   cl.metrics_small_bags, cl.metrics_large_bags, cl.metrics_pounds,
                   c.photo_url, c.review_status, c.submitted_at
            FROM contributions c
            LEFT JOIN cleanups cl ON cl.id = c.cleanup_id
            WHERE c.team_event_id = :id
            ORDER BY c.submitted_at DESC
        """),
        {"id": str(team_event_id)},
    )
    return [
        {
            "id": str(r.id),
            "user_id": str(r.user_id) if r.user_id else None,
            "team_id": str(r.team_event_team_id) if r.team_event_team_id else None,
            "contribution_type": r.contribution_type,
            "value": float(r.value) if r.value is not None else None,
            "small_bags": r.metrics_small_bags,
            "large_bags": r.metrics_large_bags,
            "pounds": float(r.metrics_pounds) if r.metrics_pounds is not None else None,
            "photo_url": r.photo_url,
            "review_status": r.review_status,
            "created_at": r.submitted_at,
        }
        for r in result.fetchall()
    ]


@router.patch("/{team_event_id}/submissions/{contribution_id}")
async def patch_submission(
    team_event_id: UUID, contribution_id: UUID, payload: PatchSubmissionRequest, db: AsyncSession = Depends(get_db)
):
    await _get_event_or_404(db, team_event_id)
    if not await _can_manage_team_event(db, team_event_id, payload.requesting_user_id):
        raise HTTPException(status_code=403, detail="Not authorized to review this event's submissions")

    prior_result = await db.execute(
        text("""
            SELECT c.cleanup_id, cl.metrics_small_bags AS small_bags, cl.metrics_large_bags AS large_bags,
                   cl.metrics_pounds AS pounds, c.value, c.review_status
            FROM contributions c
            LEFT JOIN cleanups cl ON cl.id = c.cleanup_id
            WHERE c.id = :id AND c.team_event_id = :team_event_id
        """),
        {"id": str(contribution_id), "team_event_id": str(team_event_id)},
    )
    prior = prior_result.fetchone()
    if not prior:
        raise HTTPException(status_code=404, detail="Submission not found for this event")

    fields = payload.model_dump(exclude={"requesting_user_id"}, exclude_none=True)
    if not fields:
        return {"updated": False}

    contribution_fields = {k: v for k, v in fields.items() if k in ("value", "review_status")}
    cleanup_fields = {
        {"small_bags": "metrics_small_bags", "large_bags": "metrics_large_bags", "pounds": "metrics_pounds"}[k]: v
        for k, v in fields.items()
        if k in ("small_bags", "large_bags", "pounds")
    }

    if contribution_fields:
        set_clause = ", ".join(f"{k} = :{k}" for k in contribution_fields)
        await db.execute(
            text(f"UPDATE contributions SET {set_clause} WHERE id = :id"),
            {**contribution_fields, "id": str(contribution_id)},
        )
    if cleanup_fields:
        if not prior.cleanup_id:
            raise HTTPException(status_code=400, detail="This submission has no underlying cleanup to correct")
        set_clause = ", ".join(f"{k} = :{k}" for k in cleanup_fields)
        await db.execute(
            text(f"UPDATE cleanups SET {set_clause} WHERE id = :id"),
            {**cleanup_fields, "id": str(prior.cleanup_id)},
        )
    await db.execute(
        text("""
            INSERT INTO team_event_submission_edits
                (contribution_id, edited_by,
                 previous_small_bags, previous_large_bags, previous_pounds, previous_value, previous_review_status,
                 new_small_bags, new_large_bags, new_pounds, new_value, new_review_status)
            VALUES
                (:contribution_id, :edited_by,
                 :prev_small_bags, :prev_large_bags, :prev_pounds, :prev_value, :prev_review_status,
                 :new_small_bags, :new_large_bags, :new_pounds, :new_value, :new_review_status)
        """),
        {
            "contribution_id": str(contribution_id),
            "edited_by": str(payload.requesting_user_id),
            "prev_small_bags": prior.small_bags,
            "prev_large_bags": prior.large_bags,
            "prev_pounds": prior.pounds,
            "prev_value": prior.value,
            "prev_review_status": prior.review_status,
            "new_small_bags": fields.get("small_bags", prior.small_bags),
            "new_large_bags": fields.get("large_bags", prior.large_bags),
            "new_pounds": fields.get("pounds", prior.pounds),
            "new_value": fields.get("value", prior.value),
            "new_review_status": fields.get("review_status", prior.review_status),
        },
    )
    await db.commit()
    return {"updated": True}
